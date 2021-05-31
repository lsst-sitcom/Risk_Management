import pandas as pd
from datetime import datetime, timedelta, timezone
import getpass
import openpyxl
import pytz
from jira import JIRA
#import xlsxwriter

present = datetime.now()
trigger_months = present + timedelta(weeks=26)
aztz = pytz.timezone('US/Arizona')

# this is the Excel filename that the tables will be written to
output_file = 'output.xlsx'

# prompt for Jira username
username = input("Please enter your Jira username: ")

# prompt for Jira password
password = getpass.getpass(prompt="Please enter your Jira password: ")

# prompt for how many days the report should go back
days = 30
daysIn = input("Please enter the days in the past the report should analyze (default 30): ")
try:
	days = int(daysIn)
except ValueError:
	print("Input is not an integer, defaulting to 30.\n")
	days = 30
	input("Press Enter to continue.")

# set up the JIRA issue query. This can be changed if, for example, CAM/DOE risks should be included. This is JQL and will be sent to JIRA.
components = " AND component not in ('Camera', 'DOE Funded Commissioning', 'DOE Funded Operations', 'Other', 'Operations') "
query = "project=RM AND status in ('Active Risk/Opportunity', 'Proposed', 'Retired', 'Realized')" + components + "ORDER BY cf[13108]"
query_mit = "project=RM AND issueType = RM-Handling AND status != 'Unplanned'" + components + "ORDER BY cf[13114]"

def connect():
	# establish JIRA connection

	server = "https://jira.lsstcorp.org"
	auth_inf = (username,password) # get auth info from lines 1 and 2 of file
	#try:
	jira = JIRA(server=server,basic_auth=auth_inf)
	#except:
	#	print("ERROR: Jira authentication failed. Have you provided the correct username and password?")
	#	return

	# query JIRA with the above query
	print("\nQuerying JIRA for Risk/Opportunities: \n\n" + query)
	issues = jira.search_issues(query,maxResults=None)

	# query JIRA with the above query_mit
	print("\nQuerying JIRA for Mitigations: \n\n" + query_mit)
	issues_mit= jira.search_issues(query_mit,maxResults=None)

	analyze(issues,issues_mit,jira)

def analyze(issues,issues_mit,jira):
	print("\nReturned " + str(len(issues)) + " risk/opportunity issues.")
	print("\nReturned " + str(len(issues_mit)) + " mitigation issues.")

	print("Cleaning data...")
	data = []
	data_mit = []

	# get mean probabilities depending on "current probability" range
	mean_prob_lookup = {'0%-1%':0.005,
						'0%-5%':0.025,
						'5%-10%':0.075,
						'10%-25%':0.17,
						'25%-50%':0.37,
						'50%-75%':0.63,
						'75%-100%':0.88}

	# iterate through the list of returned issues, pulling out the relevant data, one issue per row
	# each relevant field is stored as a dict entry, in the "data" list
	for i in range(len(issues)):
		#Check if not RM-Risk since others don't have PWE after mitigation and Is Parent field
		iType = issues[i].fields.issuetype.name
		id_p = issues[i].key
		if iType != 'RM-Risk':
			PWEa = 0
			isParent = 'False'
			monthsAfterMit = 0
			scheduleRecoverable = "N/A"
		else:
			PWEa = issues[i].fields.customfield_14802
			isParent = issues[i].fields.customfield_15104.value
			monthsAfterMit = issues[i].fields.customfield_14805
			scheduleRecoverable = issues[i].fields.customfield_13405
			if isParent == 'True':
				id_p = str(issues[i].key) + '(P)'

		data.append({'JIRA ID':issues[i].key,
			'Parent':isParent,
			'JIRA ID (Parent)':id_p,
			'Summary':issues[i].fields.summary,
			'issue_type':issues[i].fields.issuetype.name,
			'description':issues[i].fields.description,
			'Status':issues[i].fields.status.name,
			'Subsystem':issues[i].fields.components[0].name,
			'Probability':issues[i].fields.customfield_13200.value,
			'mean_probability':float(mean_prob_lookup[issues[i].fields.customfield_13200.value]),
			'Nonlabor Cost':issues[i].fields.customfield_13404,
			'Labor Cost':issues[i].fields.customfield_13606,
			'Model':str(issues[i].fields.customfield_13107),
			'Schedule Recoverable':scheduleRecoverable,
			'Trigger Date':issues[i].fields.customfield_13108,
			'Review Date':issues[i].fields.updated,
			'Random Period':issues[i].fields.customfield_13111,
			'Probability Weighted Exposure ($K)':issues[i].fields.customfield_14809,
			'Probability Weighted Exposure ($K) After Mitigation':PWEa,
			'WBS':issues[i].fields.customfield_12914,
			'Proposed Management Response':issues[i].fields.customfield_13501,
			'score':int(float(issues[i].fields.customfield_13900)),
			'Mitigated By':issues[i].fields.issuelinks,
			'Expected Months After Mitigation':monthsAfterMit
			})

	# iterate through the list of returned mitigation issues, pulling out the relevant data, one issue per row
	# each relevant field is stored as a dict entry, in the "data_mit" list
	for i in range(len(issues_mit)):
		data_mit.append({'JIRA ID':issues_mit[i].key,
			'Summary':issues_mit[i].fields.summary,
			'Issue Type':issues_mit[i].fields.issuetype.name,
			'Description':issues_mit[i].fields.description,
			'Status':issues_mit[i].fields.status.name,
			'Subsystem':issues_mit[i].fields.components[0].name,
			'Probability After Mitigation':issues_mit[i].fields.customfield_14504,
			'Current Handling Status':issues_mit[i].fields.customfield_13115,
			'Expected Dollars (K$)':issues_mit[i].fields.customfield_13404,
			'Expected (months)':issues_mit[i].fields.customfield_13406,
			'FTEs Required':issues_mit[i].fields.customfield_13407,
			'Anticipated Completion Date':issues_mit[i].fields.customfield_13114,
			'Review Date':issues_mit[i].fields.updated,
			'Risks Mitigated':issues_mit[i].fields.issuelinks
			})

	# iterate through and add "nonlabor PWE", "labor PWE" entries, mitigation keys
	for i in range(len(issues)):
		if data[i]['mean_probability'] == None or data[i]['Nonlabor Cost'] == None:
			data[i]['Non-labor PWE ($K)'] = 0
		else:
			data[i]['Non-labor PWE ($K)'] = data[i]['mean_probability'] * data[i]['Nonlabor Cost']

		if data[i]['mean_probability'] == None or data[i]['Labor Cost'] == None:
			data[i]['Labor PWE ($K)'] = 0
		else:
			data[i]['Labor PWE ($K)'] = data[i]['mean_probability'] * data[i]['Labor Cost']
		if data[i]['Mitigated By'] != None:
			keys = []
			for j in range(len(data[i]['Mitigated By'])):
				link = data[i]['Mitigated By'][j]
				if link.type.name == "Risk Mitigation":
					keys.append(link.inwardIssue.key)
			data[i]['Mitigated By'] = ', '.join(keys)

	# iterate through and add mitigated risks keys
	for i in range(len(issues_mit)):
		if data_mit[i]['Risks Mitigated'] != None:
			keys = []
			for j in range(len(data_mit[i]['Risks Mitigated'])):
				link_mit = data_mit[i]['Risks Mitigated'][j]
				if link_mit.type.name == "Risk Mitigation":
					keys.append(link_mit.outwardIssue.key)
				#keys.append(data_mit[i]['Risks Mitigated'][j].key)
			data_mit[i]['Risks Mitigated'] = ', '.join(keys)

	# convert to pandas dataframe
	df = pd.DataFrame(data)
	df_mit = pd.DataFrame(data_mit)

	# convert dates to datetime objects
	df['Trigger Date'] = pd.to_datetime(df['Trigger Date']).dt.tz_localize(None)
	df['Review Date'] = pd.to_datetime(df['Review Date']).dt.tz_localize(None)
	df_mit['Review Date'] = pd.to_datetime(df_mit['Review Date']).dt.tz_localize(None)
	df_mit['Anticipated Completion Date'] = pd.to_datetime(df_mit['Anticipated Completion Date']).dt.tz_localize(None)

	print("Generating tables...")

	risks = df.query("(issue_type == 'RM-Risk') & (Parent != 'True')") # all risks NOT parents
	#risks_all = df.query("(issue_type == 'RM-Risk')") # all risks
	active_risks = df.query("(issue_type == 'RM-Risk') & (Status == 'Active Risk/Opportunity') & (Parent != 'True')") # active child risks NOT parents
	active_risks_all = df.query("(issue_type == 'RM-Risk') & (Status == 'Active Risk/Opportunity')") # active risks
	non_active_risks = df.query("(issue_type == 'RM-Risk') & (Status != 'Active Risk/Opportunity')") # non-active risks
	opps = df.query("issue_type == 'RM-Opportunity'") # all opps
	active_opps = df.query("(issue_type == 'RM-Opportunity') & (Status == 'Active Risk/Opportunity')") # active opps
	#parent_risks = df.query("(issue_type == 'RM-Risk') & (Parent == 'True')") # all parent risks

	# risk overview and risk exposure tables
	table_1a = pd.pivot_table(risks,index='Subsystem',columns='Status',values='Probability Weighted Exposure ($K)',aggfunc=len,fill_value=0,margins=False)
	table_1b = pd.pivot_table(active_risks,index='Subsystem',values=['Probability Weighted Exposure ($K) After Mitigation','Probability Weighted Exposure ($K)','Non-labor PWE ($K)','Labor PWE ($K)'],aggfunc=sum,margins=False)

	# opportunity overview and exposure
	table_2a = pd.pivot_table(opps,index='Subsystem',columns='Status',values='Probability Weighted Exposure ($K)',aggfunc=len,fill_value=0,margins=False,dropna=False)
	table_2b = pd.pivot_table(opps.query("Status=='Active Risk/Opportunity'"),index='Subsystem',values=['Probability Weighted Exposure ($K)'],aggfunc=sum,margins=False,dropna=False)

	# top 10 risks based on PWE
	sorted_risks = active_risks_all.sort_values(by='Probability Weighted Exposure ($K)',ascending=False)
	table_3 = sorted_risks[0:10][['JIRA ID (Parent)','Subsystem','WBS','Summary','Probability Weighted Exposure ($K)','Probability Weighted Exposure ($K) After Mitigation']]

	# top 10 opps based on PWE
	sorted_opps = active_opps.sort_values(by='Probability Weighted Exposure ($K)',ascending=False)
	table_4 = sorted_opps[0:10][['JIRA ID','Subsystem','WBS','Summary','Probability Weighted Exposure ($K)']]

	# critical risks (risk score in the red zone of matrix, ie >= 15)
	table_6 = active_risks_all.query("score >= 15").sort_values(by='Probability Weighted Exposure ($K)',ascending=False)[:][['JIRA ID (Parent)','Subsystem','WBS','Summary','Probability Weighted Exposure ($K)','Probability Weighted Exposure ($K) After Mitigation','Proposed Management Response']]

	# trigger date outlook (within next 6 months)
	table_7 = active_risks_all[active_risks_all['Trigger Date'] < trigger_months].query('Model == "Trigger date"').sort_values(by='Trigger Date',ascending=True)[:][['Subsystem','JIRA ID (Parent)','Summary','WBS','Trigger Date','Probability','Nonlabor Cost','Labor Cost','Probability Weighted Exposure ($K)','Probability Weighted Exposure ($K) After Mitigation']]
	table_7.append(table_6.sum(numeric_only=True), ignore_index=True)

	# get risks reviewed in last 30 days
	last_reviewed = df[df['Review Date'] > (datetime.now() - timedelta(days=days))].sort_values(by='Review Date',ascending=True)

	c_table = []
	print("Done!")

	print(str(len(last_reviewed)) + " issues have been reviewed in the last " + str(days) + " days.")
	print("Querying issue comments...")

	# iterate through each risk reviewed in the last 30 days
	for i in last_reviewed.index:
		curr = last_reviewed.loc[i,'JIRA ID'] # get ID
		comments = jira.comments(jira.issue(curr)) # query JIRA for the issue's comments
		for comment in comments: # iterate through comments
			c_date = pd.to_datetime(comment.created).tz_localize(None) # convert to localized datetime
			#if (pd.to_datetime(comment.created) > (datetime.now() - timedelta(days=30))):
			if (c_date > (datetime.now() - timedelta(days = days))): # if comment is in the last 30 days
				# append to a list with all the comments
				c_table.append({'Subsystem':last_reviewed.loc[i,'Subsystem'],'ID':curr,'Summary':last_reviewed.loc[i,'Summary'],'Date':c_date,'User':comment.author.displayName,'Comment':comment.body})

	print("Done!")

	# convert comments list to a pandas dataframe
	c_table = pd.DataFrame(c_table)
	# create pivot table
	try:
		table_8 = pd.pivot_table(c_table,index=['Subsystem','ID','Summary','Date','User'],values=['Comment'],aggfunc='sum')
	except KeyError:
		table_8 = c_table
		print("Error generating comment pivot table. This can happen if there were no comments made in the selected period.")

	print("Getting changelogs...")
	change_table = []
	for i in last_reviewed.index:
		curr = last_reviewed.loc[i,'JIRA ID'] # get ID
		ci = jira.issue(curr,expand='changelog') # query JIRA for the issue's comments
		for history in ci.changelog.histories: # iterate through comments
			for item in history.items:
				c_date = pd.to_datetime(history.created).tz_localize(None) # convert to localized datetime
				if (c_date > (datetime.now() - timedelta(days = days))): # if change is in the last 30 days
					change_table.append({'ID':curr,'Date':c_date,'Author':history.author.displayName,'Field':item.field,'From':item.fromString,'To':item.toString})
	changes = pd.DataFrame(change_table)
	#print(changes)
	table_9 = pd.pivot_table(changes,index=['ID','Date','Author','Field','From','To'],aggfunc='first',fill_value = ' ')

	#df['Review Date'] = df['Review Date'].dt.tz_localize(None) # remove timezone of datetime since Excel does not support it
	#df['Trigger Date'] = df['Trigger Date'].dt.tz_localize(None)

	book = openpyxl.load_workbook('R&O Report Template.xlsx')
	filename = "Risk Report Data for " + str(datetime.now().year) + "-" + str(datetime.now().month) + ".xlsx"
	writer = pd.ExcelWriter(filename, datetime_format='YYYY-MM-DD HH:MM:SS', engine='openpyxl')
	writer.book = book
	writer.sheets = dict((ws.title,ws) for ws in book.worksheets)

	table_1a.to_excel(writer,'Table 1',startrow=11,startcol=11)
	table_1b.to_excel(writer,'Table 1',startrow=2,startcol=11,header=False)
	table_2b.to_excel(writer,'Table 2',startrow=2,startcol=10,header=False)
	table_2a.to_excel(writer,'Table 2',startrow=11,startcol=10)
	table_3.to_excel(writer,'Table 3',startrow=3,startcol=1,index=False,header=False)
	table_4.to_excel(writer,'Table 4',startrow=3,startcol=1,index=False,header=False)
	table_6.to_excel(writer,'Table 6',startrow=1,index=False,header=False)
	table_7.to_excel(writer,'Table 7',startrow=1,index=False,header=False)
	table_8.to_excel(writer,'Table 8',startrow=1,header=False)
	table_9.to_excel(writer,'Table 9',startrow=1,na_rep = ' ',header=False)
	active_risks_all.to_excel(writer,'All Active Risks',startrow=0,index=False)
	non_active_risks.to_excel(writer,'All Non-Active Risks',startrow=0,index=False)
	df_mit.to_excel(writer,'All Mitigations',startrow=0,index=False)

	try:
		writer.save()
		print("Saved!")
	except PermissionError:
		print("ERROR: could not write file. Please ensure the Excel file is not already open.")

	wb = openpyxl.load_workbook(filename)
	date_style = openpyxl.styles.NamedStyle(name='datetime', number_format='DD/MM/YYYY')
	ws = wb["Table 7"]
	for cell in ws['E:E']:
		if cell.value != "Trigger Date":
			cell.style = date_style
	wb.save(filename=filename)


	input("\nPress Enter to close.")
	return

connect()
